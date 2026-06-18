# -*- coding: utf-8 -*-
"""
=================================================================================
 SISTEMA DE CONTROLE DE METAS DE QUILOMETRAGEM - MOTORISTAS  (v3.2)
 EPR VIA MINEIRA - COM CAIXA DE SELEÇÃO DE MOTORISTAS
=================================================================================
"""

import os
import io
import re
from datetime import datetime, date

import numpy as np
import pandas as pd
import streamlit as st
from PIL import Image, ImageOps
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# =================================================================================
# 1. CONFIGURAÇÕES GERAIS E LISTAS DE MOTORISTAS
# =================================================================================

ARQUIVO_PLANTAO = "registros_plantao.csv"
ARQUIVO_OCORRENCIAS = "registros_ocorrencias.csv"

META_IDEAL_KM = 400.0     
META_MINIMA_KM = 380.0    

MOTOR_OCR = "pytesseract"

# --- LISTA OFICIAL DE MOTORISTAS (BSO-01 E BSO-02) ---
# Organizados em ordem alfabética para facilitar a busca na caixinha
TODOS_MOTORISTAS = [
    "Claudio Roberto",
    "Edson",
    "Eduardo",
    "Elias Cruz",
    "Evander",
    "Franciele",
    "Gilsimar",
    "Hugo Leonardo",
    "Leo Junior",
    "Leticia Souza",
    "Luciano Pedro",
    "Roberto Carlos",
    "Romulo",
    "Valeria",
     "Douglas",
    "Verificar Nome" # Opção de segurança caso o CCO mande alguém novo
]

COLUNAS_PLANTAO = [
    "data_plantao", "turno", "base", "colaborador", "vtr", "km_rodados",
    "status_meta", "tempo_parado_total", "registrado_em",
]
COLUNAS_OCORRENCIAS = [
    "data_plantao", "turno", "colaborador", "numero_ocorrencia", "tempo_parado",
]

DIAS_SEMANA_PT = {
    0: "SEGUNDA-FEIRA", 1: "TERÇA-FEIRA", 2: "QUARTA-FEIRA",
    3: "QUINTA-FEIRA", 4: "SEXTA-FEIRA", 5: "SÁBADO", 6: "DOMINGO",
}

LOGO_URL = "https://eprviamineira.com.br/wp-content/uploads/2024/10/epr-via-mineira-logo.png"


st.set_page_config(
    page_title="EPR Via Mineira | Controle de Metas de KM",
    page_icon="🛣️",
    layout="wide",
    initial_sidebar_state="expanded",
)


# =================================================================================
# 2. IDENTIDADE VISUAL - CSS CUSTOMIZADO (EPR VIA MINEIRA)
# =================================================================================

def inject_custom_css():
    st.markdown(
        """
        <style>
        :root {
            --epr-azul-escuro: #0B2D4E;
            --epr-azul: #154C79;
            --epr-azul-claro: #1F6FB2;
            --epr-cinza-fundo: #F2F4F7;
            --epr-cinza-borda: #D9DEE3;
            --epr-branco: #FFFFFF;
            --epr-verde: #1E8449;
            --epr-verde-claro: #E8F8F0;
            --epr-laranja: #E67E22;
            --epr-laranja-claro: #FDF2E3;
            --epr-vermelho: #C0392B;
            --epr-vermelho-claro: #FDEDEC;
            --epr-texto: #1B2631;
        }

        .stApp { background-color: var(--epr-cinza-fundo); }
        #MainMenu {visibility: hidden;}
        footer {visibility: hidden;}

        h1, h2, h3, h4 {
            color: var(--epr-azul-escuro) !important;
            font-family: "Segoe UI", Arial, sans-serif;
        }
        p, span, label, .stMarkdown { color: var(--epr-texto); }

        [data-testid="stSidebar"] {
            background-color: var(--epr-azul-escuro);
            border-right: 1px solid var(--epr-cinza-borda);
        }
        [data-testid="stSidebar"] * { color: var(--epr-branco) !important; }
        .epr-sidebar-brand {
            text-align: center;
            padding: 0.4rem 0 1.1rem 0;
            border-bottom: 1px solid rgba(255,255,255,0.18);
            margin-bottom: 0.6rem;
        }
        .epr-sidebar-brand img { height: 36px; margin-bottom: 0.5rem; background: white; border-radius: 6px; padding: 4px 8px; }
        .epr-sidebar-brand p {
            color: var(--epr-branco) !important;
            font-weight: 800;
            font-size: 0.98rem;
            margin: 0;
            letter-spacing: 0.4px;
        }
        .epr-sidebar-brand span { color: #BFD2E3 !important; font-size: 0.74rem; }
        .epr-sidebar-footer {
            text-align: center;
            color: #8CA3B8 !important;
            font-size: 0.7rem;
            margin-top: 1.4rem;
            padding-top: 0.8rem;
            border-top: 1px solid rgba(255,255,255,0.18);
        }

        div[data-testid="stVerticalBlockBorderWrapper"] {
            background-color: var(--epr-branco);
            border: 1px solid var(--epr-cinza-borda);
            border-radius: 12px;
            box-shadow: 0 2px 8px rgba(11, 45, 78, 0.06);
        }

        [data-testid="stDataFrame"] {
            border: 1px solid var(--epr-cinza-borda);
            border-radius: 10px;
            overflow: hidden;
        }

        .stButton > button, .stDownloadButton > button {
            background-color: var(--epr-azul-escuro);
            color: white;
            border: none;
            border-radius: 8px;
            font-weight: 600;
            padding: 0.55rem 1.2rem;
            transition: background-color 0.2s ease-in-out;
        }
        .stButton > button:hover, .stDownloadButton > button:hover {
            background-color: var(--epr-azul-claro);
            color: white;
        }
        .stButton > button[kind="primary"] { background-color: var(--epr-verde); }
        .stButton > button[kind="primary"]:hover { background-color: #239B56; }

        .epr-badge {
            display: block;
            text-align: center;
            font-size: 1.0rem;
            font-weight: 700;
            letter-spacing: 0.3px;
            padding: 0.65rem 1rem;
            border-radius: 8px;
            margin: 0.6rem 0 1rem 0;
        }
        .epr-badge-success { background-color: var(--epr-verde-claro); color: var(--epr-verde); border: 1.5px solid var(--epr-verde); }
        .epr-badge-warning { background-color: var(--epr-laranja-claro); color: var(--epr-laranja); border: 1.5px solid var(--epr-laranja); }
        .epr-badge-danger { background-color: var(--epr-vermelho-claro); color: var(--epr-vermelho); border: 1.5px solid var(--epr-vermelho); }

        .epr-tag {
            display: inline-block;
            padding: 0.2rem 0.65rem;
            border-radius: 999px;
            font-size: 0.74rem;
            font-weight: 700;
            letter-spacing: 0.2px;
        }
        .epr-tag-success { background-color: var(--epr-verde-claro); color: var(--epr-verde); }
        .epr-tag-warning { background-color: var(--epr-laranja-claro); color: var(--epr-laranja); }
        .epr-tag-danger { background-color: var(--epr-vermelho-claro); color: var(--epr-vermelho); }

        .epr-card-historico {
            background-color: var(--epr-branco);
            border: 1px solid var(--epr-cinza-borda);
            border-radius: 10px;
            overflow: hidden;
            margin-bottom: 1.3rem;
        }
        .epr-bloco-titulo {
            background: linear-gradient(135deg, var(--epr-azul-escuro) 0%, var(--epr-azul) 100%);
            color: white !important;
            font-weight: 700;
            font-size: 0.98rem;
            padding: 0.7rem 1.1rem;
        }
        .epr-table { width: 100%; border-collapse: collapse; }
        .epr-table th {
            background-color: #F7F9FA;
            color: var(--epr-azul-escuro);
            text-align: left;
            font-size: 0.74rem;
            font-weight: 700;
            padding: 0.55rem 1rem;
            border-bottom: 2px solid var(--epr-cinza-borda);
            text-transform: uppercase;
            letter-spacing: 0.3px;
        }
        .epr-table td {
            padding: 0.55rem 1rem;
            font-size: 0.88rem;
            border-bottom: 1px solid #EDEFF2;
            color: var(--epr-texto);
        }

        .epr-footer {
            text-align: center;
            color: #8C98A4;
            font-size: 0.8rem;
            margin-top: 2rem;
            padding-top: 1rem;
            border-top: 1px solid var(--epr-cinza-borda);
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_sidebar_brand():
    st.sidebar.markdown(
        f"""
        <div class="epr-sidebar-brand">
            <img src="{LOGO_URL}" onerror="this.style.display='none'" />
            <p>EPR VIA MINEIRA</p>
            <span>Controle de Metas de KM</span>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_sidebar_footer():
    st.sidebar.markdown(
        '<div class="epr-sidebar-footer">Sistema interno &middot; uso operacional</div>',
        unsafe_allow_html=True,
    )


# =================================================================================
# 3. FUNÇÕES UTILITÁRIAS (regras de negócio / formatação)
# =================================================================================

def calcular_status(km_rodados: float):
    km_rodados = float(km_rodados or 0)
    if km_rodados >= META_IDEAL_KM:
        return "BATIDA", "META BATIDA", "epr-badge-success"
    if km_rodados >= META_MINIMA_KM:
        return "ACEITAVEL", "META ACEITÁVEL", "epr-badge-warning"
    return "NAO_BATIDA", "META NÃO BATIDA", "epr-badge-danger"


def format_data_br(iso_date_str: str) -> str:
    try:
        return datetime.strptime(str(iso_date_str), "%Y-%m-%d").strftime("%d/%m/%Y")
    except (ValueError, TypeError):
        return str(iso_date_str)


def nome_dia_semana_pt(d: date) -> str:
    return DIAS_SEMANA_PT.get(d.weekday(), "")


def hhmm_to_minutes(valor: str) -> int:
    try:
        partes = str(valor).strip().split(":")
        horas = int(partes[0])
        minutos = int(partes[1]) if len(partes) > 1 else 0
        return horas * 60 + minutos
    except (ValueError, IndexError):
        return 0


def minutes_to_hhmm(total_minutos: int) -> str:
    total_minutos = max(0, int(total_minutos))
    horas, minutos = divmod(total_minutos, 60)
    return f"{horas:02d}:{minutos:02d}"


def somar_tempos(lista_tempos) -> str:
    total = sum(hhmm_to_minutes(t) for t in lista_tempos)
    return minutes_to_hhmm(total)


# =================================================================================
# 4. CAMADA DE DADOS
# =================================================================================

def load_plantao_df() -> pd.DataFrame:
    if os.path.exists(ARQUIVO_PLANTAO):
        df = pd.read_csv(ARQUIVO_PLANTAO, dtype=str)
        for col in COLUNAS_PLANTAO:
            if col not in df.columns:
                df[col] = ""
        return df[COLUNAS_PLANTAO]
    return pd.DataFrame(columns=COLUNAS_PLANTAO)


def load_ocorrencias_df() -> pd.DataFrame:
    if os.path.exists(ARQUIVO_OCORRENCIAS):
        df = pd.read_csv(ARQUIVO_OCORRENCIAS, dtype=str)
        for col in COLUNAS_OCORRENCIAS:
            if col not in df.columns:
                df[col] = ""
        return df[COLUNAS_OCORRENCIAS]
    return pd.DataFrame(columns=COLUNAS_OCORRENCIAS)


def salvar_plantao_atual(data_plantao_iso: str, turno: str, lista_motoristas: list) -> list:
    if not lista_motoristas:
        return ["Não há motoristas confirmados para salvar."]

    erros = []
    for idx, motorista in enumerate(lista_motoristas, start=1):
        nome = str(motorista.get("colaborador", "")).strip()
        if not nome or nome == "Verificar Nome":
            erros.append(f"Linha {idx}: selecione o nome correto do colaborador na lista.")
            continue
        status_code, _, _ = calcular_status(motorista.get("km_rodados"))
        if status_code != "BATIDA" and not motorista.get("ocorrencias"):
            erros.append(f"{nome}: meta não batida/aceitável precisa de ao menos 1 ocorrência registrada.")

    if erros:
        return erros

    linhas_plantao, lines_ocorrencias = [], []
    for motorista in lista_motoristas:
        status_code, _, _ = calcular_status(motorista.get("km_rodados"))
        tempos = [oc["tempo"] for oc in motorista.get("ocorrencias", [])]
        linhas_plantao.append({
            "data_plantao": data_plantao_iso,
            "turno": turno,
            "base": str(motorista.get("base", "")).strip() or "-",
            "colaborador": str(motorista.get("colaborador", "")).strip(),
            "vtr": str(motorista.get("vtr", "")).strip() or "-",
            "km_rodados": round(float(motorista.get("km_rodados") or 0), 1),
            "status_meta": status_code,
            "tempo_parado_total": somar_tempos(tempos),
            "registrado_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })
        for oc in motorista.get("ocorrencias", []):
            lines_ocorrencias.append({
                "data_plantao": data_plantao_iso,
                "turno": turno,
                "colaborador": str(motorista.get("colaborador", "")).strip(),
                "numero_ocorrencia": oc["numero"],
                "tempo_parado": oc["tempo"],
            })

    df_plantao = load_plantao_df()
    chave_atual = (df_plantao["data_plantao"] == data_plantao_iso) & (df_plantao["turno"] == turno)
    df_plantao = df_plantao[~chave_atual]
    df_plantao = pd.concat([df_plantao, pd.DataFrame(linhas_plantao, columns=COLUNAS_PLANTAO)], ignore_index=True)
    df_plantao.to_csv(ARQUIVO_PLANTAO, index=False)

    df_ocorrencias = load_ocorrencias_df()
    chave_oc_atual = (df_ocorrencias["data_plantao"] == data_plantao_iso) & (df_ocorrencias["turno"] == turno)
    df_ocorrencias = df_ocorrencias[~chave_oc_atual]
    df_ocorrencias = pd.concat(
        [df_ocorrencias, pd.DataFrame(lines_ocorrencias, columns=COLUNAS_OCORRENCIAS)], ignore_index=True
    )
    df_ocorrencias.to_csv(ARQUIVO_OCORRENCIAS, index=False)

    return []


# =================================================================================
# 5. INTELIGÊNCIA OCR ATUALIZADA (FILTRO DIRETO PARA PLANILHA REAL DO CCO)
# =================================================================================

def preprocessar_imagem_para_ocr(image: Image.Image) -> Image.Image:
    imagem = image.convert("L")
    imagem = ImageOps.autocontrast(imagem)
    largura, altura = imagem.size
    if largura < 1500:
        fator = 1500 / largura
        imagem = imagem.resize((int(largura * fator), int(altura * fator)), Image.LANCZOS)
    return imagem


def interpretar_texto_completo_real(linhas_texto: list) -> list:
    """
    Motor inteligente feito sob medida para ler o print real do CCO da EPR.
    """
    resultados_filtrados = []
    
    padrao_base = re.compile(r"(BSO-0[12])", re.IGNORECASE)
    padrao_vtr = re.compile(r"\b(T\d{2})\b", re.IGNORECASE)
    padrao_numeros = re.compile(r"\b(\d{2,3})\b")

    base_atual = "BSO-01"

    for linha in linhas_texto:
        match_base = padrao_base.search(linha)
        if match_base:
            base_atual = match_base.group(1).upper()

        if "BSO-03" in linha.upper() or "BSO-04" in linha.upper() or "BSO-05" in linha.upper():
            continue

        match_vtr = padrao_vtr.search(linha)
        if match_vtr:
            vtr_encontrada = match_vtr.group(1).upper()
            
            numeros_linha = padrao_numeros.findall(linha)
            km_rodados = 0.0
            
            candidatos_km = [float(n) for x in numeros_linha if 50 <= (float(x)) <= 500]
            if candidatos_km:
                km_rodados = candidatos_km[0] 

            colaborador = re.sub(r"[^A-Za-zÀ-ÿ\s]", " ", linha)
            colaborador = re.sub(r"\bBSO\b|\bT\d{2}\b", " ", colaborador, flags=re.IGNORECASE)
            colaborador = " ".join(colaborador.split()).title()

            # Tenta combinar o nome lido com algum da nossa lista oficial
            nome_encontrado = "Verificar Nome"
            for nome_oficial in TODOS_MOTORISTAS:
                if nome_oficial.lower() in colaborador.lower():
                    nome_encontrado = nome_oficial
                    break

            resultados_filtrados.append({
                "base": base_atual,
                "colaborador": nome_encontrado,
                "vtr": vtr_encontrada,
                "km_rodados": km_rodados
            })

    return resultados_filtrados


def process_ocr_multi(image: Image.Image) -> list:
    imagem_processada = preprocessar_imagem_para_ocr(image)
    try:
        import pytesseract
        from pytesseract import Output
        dados = pytesseract.image_to_data(imagem_processada, lang="por", output_type=Output.DATAFRAME, config="--psm 6")
        dados = dados.dropna(subset=["text"])
        dados["text"] = dados["text"].astype(str).str.strip()
        dados = dados[dados["text"] != ""]

        linhas_texto = []
        for _, grupo in dados.groupby(["block_num", "par_num", "line_num"], sort=False):
            grupo_ordenado = grupo.sort_values("left")
            linhas_texto.append(" ".join(grupo_ordenado["text"].tolist()))
            
        return interpretar_texto_completo_real(linhas_texto)

    except Exception:
        return []


# =================================================================================
# 6. EXPORTAÇÃO EXCEL ATUALIZADA
# =================================================================================

def export_consolidated_excel(df_plantao: pd.DataFrame, df_ocorrencias: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório Consolidado"

    azul, cinza_claro = "0B2D4E", "F2F4F7"
    verde_claro, laranja_claro, vermelho_claro = "E8F8F0", "FDF2E3", "FDEDEC"
    branco = "FFFFFF"

    fonte_titulo = Font(bold=True, color=branco, size=12)
    fonte_header = Font(bold=True, color=azul, size=10)
    fonte_normal = Font(size=10)
    fonte_status = Font(bold=True, size=10)

    fill_titulo = PatternFill("solid", fgColor=azul)
    fill_header = PatternFill("solid", fgColor=cinza_claro)
    fill_verde = PatternFill("solid", fgColor=verde_claro)
    fill_laranja = PatternFill("solid", fgColor=laranja_claro)
    fill_vermelho = PatternFill("solid", fgColor=vermelho_claro)

    for idx, largura in enumerate([12, 28, 14, 14, 18, 20, 18], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = largura

    rotulos_status = {"BATIDA": "META BATIDA", "ACEITAVEL": "META ACEITÁVEL", "NAO_BATIDA": "META NÃO BATIDA"}
    fills_status = {"BATIDA": fill_verde, "ACEITAVEL": fill_laranja, "NAO_BATIDA": fill_vermelho}
    colunas_cabecalho = ["Base", "Colaborador", "VTR", "KM Rodados", "Status da Meta", "Nº Ocorrência(s)", "Tempo Parado Total"]
    ordem_turno = {"DIA": 0, "NOITE": 1}

    df_plantao = df_plantao.copy()
    df_plantao["km_rodados"] = pd.to_numeric(df_plantao["km_rodados"], errors="coerce")

    chaves = df_plantao[["data_plantao", "turno"]].drop_duplicates()
    chaves_ordenadas = sorted(
        chaves.itertuples(index=False, name=None),
        key=lambda k: (k[0], ordem_turno.get(k[1], 2)),
        reverse=True,
    )

    linha_atual = 1
    for data_str, turno in chaves_ordenadas:
        bloco = df_plantao[(df_plantao["data_plantao"] == data_str) & (df_plantao["turno"] == turno)]
        try:
            data_fmt = datetime.strptime(data_str, "%Y-%m-%d").date()
            titulo_bloco = f"{data_fmt.strftime('%d/%m')} {nome_dia_semana_pt(data_fmt)} · TURNO {turno}"
        except ValueError:
            titulo_bloco = f"{data_str} · TURNO {turno}"

        ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=7)
        celula_titulo = ws.cell(row=linha_atual, column=1, value=titulo_bloco)
        celula_titulo.font = fonte_titulo
        celula_titulo.fill = fill_titulo
        celula_titulo.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[linha_atual].height = 22
        linha_atual += 1

        for col_idx, titulo_col in enumerate(colunas_cabecalho, start=1):
            celula = ws.cell(row=linha_atual, column=col_idx, value=titulo_col)
            celula.font = fonte_header
            celula.fill = fill_header
            celula.alignment = Alignment(horizontal="left")
        linha_atual += 1

        for _, linha in bloco.iterrows():
            status_code = linha["status_meta"]
            
            ocorrencias_motorista = df_ocorrencias[
                (df_ocorrencias["data_plantao"] == data_str)
                & (df_ocorrencias["turno"] == turno)
                & (df_ocorrencias["colaborador"] == linha["colaborador"])
            ]
            numeros_oc = ", ".join(ocorrencias_motorista["numero_ocorrencia"].tolist()) if not ocorrencias_motorista.empty else "-"

            valores = [
                linha.get("base", "-"), 
                linha["colaborador"], 
                linha["vtr"],
                f"{float(linha['km_rodados']):.1f} km",
                rotulos_status.get(status_code, status_code),
                numeros_oc,
                linha.get("tempo_parado_total", "") or "-",
            ]
            for col_idx, valor in enumerate(valores, start=1):
                celula = ws.cell(row=linha_atual, column=col_idx, value=valor)
                celula.font = fonte_status if col_idx == 5 else fonte_normal
                if col_idx == 5:
                    celula.fill = fills_status.get(status_code, fill_header)
            linha_atual += 1

        linha_atual += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# =================================================================================
# 7. COMPONENTES DE INTERFACE REUTILIZÁVEIS
# =================================================================================

def _sincronizar_confirmacao():
    ocorrencias_existentes = {
        str(m.get("colaborador", "")).strip().lower(): m.get("ocorrencias", [])
        for m in st.session_state.get("plantao_motoristas", [])
        if m.get("colaborador")
    }
    nova_lista = []
    for linha in st.session_state.linhas_ocr:
        nome = str(linha.get("colaborador", "")).strip()
        if not nome or nome == "Verificar Nome":
            continue
        chave = nome.lower()
        nova_lista.append({
            "base": str(linha.get("base", "")).strip(),
            "colaborador": nome,
            "vtr": str(linha.get("vtr", "")).strip() or "-",
            "km_rodados": round(float(linha.get("km_rodados") or 0), 1),
            "ocorrencias": ocorrencias_existentes.get(chave, []),
        })
    st.session_state.plantao_motoristas = nova_lista


def _callback_adicionar_ocorrencia(indice: int):
    chave_numero = f"novo_num_{indice}"
    chave_tempo = f"novo_tempo_{indice}"
    numero = st.session_state.get(chave_numero, "").strip()
    tempo = st.session_state.get(chave_tempo, "").strip()

    if not numero:
        st.session_state[f"erro_oc_{indice}"] = "Informe o número da ocorrência."
        return
    if not re.match(r"^\d{1,3}:\d{2}$", tempo):
        st.session_state[f"erro_oc_{indice}"] = "Tempo parado deve estar no formato HH:MM (ex: 01:30)."
        return

    st.session_state[f"erro_oc_{indice}"] = ""
    st.session_state.plantao_motoristas[indice].setdefault("ocorrencias", []).append(
        {"numero": numero, "tempo": tempo}
    )
    st.session_state[chave_numero] = ""
    st.session_state[chave_tempo] = ""


def _renderizar_card_motorista(i: int, motorista: dict):
    with st.container(border=True):
        col_nome, col_vtr, col_km, col_rm = st.columns([3, 1.3, 1.3, 0.7])
        with col_nome:
            st.markdown(f"**{motorista['colaborador']}** | {motorista.get('base', '-')}")
        with col_vtr:
            st.caption(f"VTR: {motorista['vtr']}")
        with col_km:
            st.caption(f"{float(motorista['km_rodados']):.1f} km")
        with col_rm:
            if st.button("🗑️", key=f"rm_{i}", use_container_width=True):
                st.session_state.plantao_motoristas.pop(i)
                st.rerun()

        status_code, status_label, classe_badge = calcular_status(motorista["km_rodados"])
        st.markdown(
            f'<div class="epr-badge {classe_badge}">{status_label} &nbsp;|&nbsp; '
            f'{float(motorista["km_rodados"]):.1f} km</div>',
            unsafe_allow_html=True,
        )

        if status_code != "BATIDA":
            st.markdown("**⚠️ Ocorrências deste motorista no turno (é possível empilhar mais de uma):**")
            ocorrencias = motorista.setdefault("ocorrencias", [])

            if ocorrencias:
                for j, oc in enumerate(ocorrencias):
                    col_a, col_b, col_c = st.columns([2, 1.4, 0.6])
                    col_a.write(f"📌 Ocorrência nº {oc['numero']}")
                    col_b.write(f"⏱️ {oc['tempo']}")
                    if col_c.button("🗑️", key=f"rm_oc_{i}_{j}"):
                        ocorrencias.pop(j)
                        st.rerun()
            else:
                st.caption("Nenhuma ocorrência adicionada ainda.")

            col_num, col_tempo, col_add = st.columns([2, 1.3, 0.9])
            with col_num:
                st.text_input("Nº Ocorrência (ex: oc 170)", key=f"novo_num_{i}")
            with col_tempo:
                st.text_input("Tempo Parado (HH:MM)", key=f"novo_tempo_{i}", placeholder="01:30")
            with col_add:
                st.write("")
                st.button(
                    "➕ Adicionar", key=f"add_oc_{i}",
                    on_click=_callback_adicionar_ocorrencia, args=(i,),
                    use_container_width=True,
                )

            erro_oc = st.session_state.get(f"erro_oc_{i}", "")
            if erro_oc:
                st.error(erro_oc)

            total_parado = somar_tempos([oc["tempo"] for oc in ocorrencias])
            st.markdown(f"**🕒 TOTAL PARADO: {total_parado}**")


# =================================================================================
# 8. PÁGINAS DA APLICAÇÃO
# =================================================================================

def pagina_inicio():
    st.markdown("## 🏠 Painel Geral")
    df_plantao = load_plantao_df()

    if df_plantao.empty:
        st.info("Ainda não há registros. Comece lançando o primeiro plantão em **📝 Lançar Plantão Diário**.")
        return

    df_plantao["km_rodados"] = pd.to_numeric(df_plantao["km_rodados"], errors="coerce")
    total_registros = len(df_plantao)
    total_motoristas = df_plantao["colaborador"].nunique()
    contagem_status = df_plantao["status_meta"].value_counts()
    pct_batida = 100 * contagem_status.get("BATIDA", 0) / (total_registros or 1)
    pct_nao_batida = 100 * contagem_status.get("NAO_BATIDA", 0) / (total_registros or 1)

    col1, col2, col3, col4 = st.columns(4)
    with col1, st.container(border=True):
        st.metric("Plantões Registrados", total_registros)
    with col2, st.container(border=True):
        st.metric("Motoristas Únicos", total_motoristas)
    with col3, st.container(border=True):
        st.metric("🟢 % Meta Batida", f"{pct_batida:.0f}%")
    with col4, st.container(border=True):
        st.metric("🔴 % Meta Não Batida", f"{pct_nao_batida:.0f}%")

    st.write("")
    col_graf, col_tabela = st.columns([1, 1.5], gap="large")
    with col_graf, st.container(border=True):
        st.markdown("#### Distribuição de Status")
        dados_grafico = pd.DataFrame({
            "Status": ["Meta Batida", "Meta Aceitável", "Meta Não Batida"],
            "Quantidade": [
                contagem_status.get("BATIDA", 0),
                contagem_status.get("ACEITAVEL", 0),
                contagem_status.get("NAO_BATIDA", 0),
            ],
        }).set_index("Status")
        st.bar_chart(dados_grafico)

    with col_tabela, st.container(border=True):
        st.markdown("#### Últimos Lançamentos")
        df_recentes = df_plantao.sort_values("registrado_em", ascending=False).head(8).copy()
        df_recentes["data_plantao"] = df_recentes["data_plantao"].apply(format_data_br)
        df_recentes["status_meta"] = df_recentes["status_meta"].map(
            {"BATIDA": "✅ Batida", "ACEITAVEL": "🟠 Aceitável", "NAO_BATIDA": "❌ Não Batida"}
        )
        st.dataframe(
            df_recentes[["data_plantao", "turno", "base", "colaborador", "vtr", "km_rodados", "status_meta"]],
            use_container_width=True, hide_index=True,
        )


def pagina_lancar_plantao():
    st.markdown("## 📝 Lançar Plantão Diário")

    if "linhas_ocr" not in st.session_state:
        st.session_state.linhas_ocr = []
    if "plantao_motoristas" not in st.session_state:
        st.session_state.plantao_motoristas = []

    col_data, col_turno = st.columns(2)
    with col_data:
        data_plantao = st.date_input("Data do Plantão", value=date.today(), format="DD/MM/YYYY")
    with col_turno:
        turno = st.radio("Turno", options=["DIA", "NOITE"], horizontal=True)

    st.caption(f"Meta Ideal: **{META_IDEAL_KM:.0f} km** &nbsp;|&nbsp; Mínimo Aceitável: **{META_MINIMA_KM:.0f} km**")

    with st.container(border=True):
        st.markdown("### 📥 Importar Tabela do CCO (OCR)")
        st.caption(
            "Arraste o print completo da tabela enviado pelo CCO no WhatsApp. O sistema filtrará apenas BSO-01 e BSO-02 automaticamente."
        )

        col_up, col_prev = st.columns([1, 1])
        with col_up:
            arquivo = st.file_uploader("Print do CCO", type=["png", "jpg", "jpeg"], key="upload_plantao")
            processar = st.button(
                "🔍 Processar Imagem (OCR)", disabled=arquivo is None, use_container_width=True
            )
        imagem = None
        with col_prev:
            if arquivo is not None:
                try:
                    imagem = Image.open(arquivo)
                    st.image(imagem, use_container_width=True, caption="Pré-visualização")
                except Exception:
                    st.error("Não foi possível abrir essa imagem.")

        if processar and imagem is not None:
            with st.spinner("Lendo e filtrando a tabela do CCO..."):
                novas_linhas = process_ocr_multi(imagem)
            if novas_linhas:
                st.session_state.linhas_ocr.extend(novas_linhas)
                if "editor_plantao" in st.session_state:
                    del st.session_state["editor_plantao"]
                st.success(f"{len(novas_linhas)} linha(s) de BSO-01/BSO-02 mapeada(s) com sucesso!")
            else:
                st.warning(
                    "Não encontramos dados automáticos nas linhas. Use a tabela limpa abaixo para digitar."
                )
            st.rerun()

        st.markdown("##### Tabela extraída — corrija o que for necessário")
        
        if not st.session_state.linhas_ocr:
            linhas_iniciais = [{"base": "BSO-01", "colaborador": None, "vtr": "", "km_rodados": 0.0} for _ in range(3)]
        else:
            linhas_iniciais = st.session_state.linhas_ocr
            
        df_base = pd.DataFrame(linhas_iniciais)
        
        # A MÁGICA ACONTECE AQUI: A coluna de Colaborador agora é uma Selectbox
        df_editado = st.data_editor(
            df_base,
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
            column_config={
                "base": st.column_config.SelectboxColumn(
                    "Base (BSO)", 
                    options=["BSO-01", "BSO-02"], 
                    required=True
                ),
                "colaborador": st.column_config.SelectboxColumn(
                    "Colaborador", 
                    options=TODOS_MOTORISTAS, 
                    required=True
                ),
                "vtr": st.column_config.TextColumn("VTR"),
                "km_rodados": st.column_config.NumberColumn(
                    "KM Rodados", min_value=0.0, step=0.5, format="%.1f"
                ),
            },
        )

        if st.button("✅ Confirmar Lista para Validação de Metas", type="primary", use_container_width=True):
            st.session_state.linhas_ocr = df_editado.fillna({"base": "BSO-01", "colaborador": "", "vtr": "", "km_rodados": 0.0}).to_dict("records")
            _sincronizar_confirmacao()
            st.rerun()

    lista = st.session_state.plantao_motoristas
    if not lista:
        st.info("Importe e confirme a lista de motoristas acima para liberar a validação de metas e ocorrências.")
        return

    st.write("")
    st.markdown("### ✅ Validação de Metas e Ocorrências")
    contagem = {"BATIDA": 0, "ACEITAVEL": 0, "NAO_BATIDA": 0}
    for m in lista:
        status_code, _, _ = calcular_status(m.get("km_rodados"))
        contagem[status_code] += 1
    c1, c2, c3 = st.columns(3)
    c1.metric("🟢 Batida", contagem["BATIDA"])
    c2.metric("🟠 Aceitável", contagem["ACEITAVEL"])
    c3.metric("🔴 Não Batida", contagem["NAO_BATIDA"])

    for i in range(len(st.session_state.plantao_motoristas)):
        _renderizar_card_motorista(i, st.session_state.plantao_motoristas[i])

    st.write("")
    with st.container(border=True):
        col_a, col_b, col_c = st.columns([1, 2, 1])
        with col_b:
            if st.button("💾 Salvar Plantão", type="primary", use_container_width=True):
                erros = salvar_plantao_atual(
                    data_plantao.isoformat(), turno, st.session_state.plantao_motoristas
                )
                if erros:
                    for erro in erros:
                        st.error(erro)
                else:
                    qtd = len(st.session_state.plantao_motoristas)
                    st.success(
                        f"Plantão de {format_data_br(data_plantao.isoformat())} ({turno}) salvo!"
                    )
                    st.session_state.plantao_motoristas = []
                    st.session_state.linhas_ocr = []
                    if "editor_plantao" in st.session_state:
                        del st.session_state["editor_plantao"]
                    st.rerun()


def _renderizar_bloco_historico(data_str: str, turno: str, bloco_df: pd.DataFrame, df_ocorrencias: pd.DataFrame):
    try:
        data_fmt = datetime.strptime(data_str, "%Y-%m-%d").date()
        titulo = f"{data_fmt.strftime('%d/%m')} {nome_dia_semana_pt(data_fmt)} · TURNO {turno}"
    except ValueError:
        titulo = f"{data_str} · TURNO {turno}"

    icone_turno = "☀️" if turno == "DIA" else "🌙"
    rotulos_status = {
        "BATIDA": ("META BATIDA", "epr-tag-success"),
        "ACEITAVEL": ("META ACEITÁVEL", "epr-tag-warning"),
        "NAO_BATIDA": ("META NÃO BATIDA", "epr-tag-danger"),
    }

    linhas_html = ""
    for _, linha in bloco_df.iterrows():
        status_code = linha["status_meta"]
        status_label, classe_tag = rotulos_status.get(status_code, (status_code, ""))

        ocorrencias_motorista = df_ocorrencias[
            (df_ocorrencias["data_plantao"] == data_str)
            & (df_ocorrencias["turno"] == turno)
            & (df_ocorrencias["colaborador"] == linha["colaborador"])
        ]
        
        numeros_oc = ", ".join(ocorrencias_motorista["numero_ocorrencia"].tolist())
        if not numeros_oc:
            numeros_oc = "-"

        linhas_html += f"""
        <tr>
            <td>{linha.get('base', '')}</td>
            <td>{linha['colaborador']}</td>
            <td>{linha.get('vtr', '') or '-'}</td>
            <td>{float(linha['km_rodados']):.1f} km</td>
            <td><span class="epr-tag {classe_tag}">{status_label}</span></td>
            <td>{numeros_oc}</td>
            <td>{linha.get('tempo_parado_total', '') or '-'}</td>
        </tr>
        """

    st.markdown(
        f"""
        <div class="epr-card-historico">
            <div class="epr-bloco-titulo">{icone_turno} {titulo}</div>
            <table class="epr-table">
                <thead>
                    <tr>
                        <th>Base</th><th>Colaborador</th><th>VTR</th>
                        <th>KM Rodados</th><th>Status</th><th>Nº Ocorrência(s)</th><th>Tempo Parado Total</th>
                    </tr>
                </thead>
                <tbody>
                    {linhas_html}
                </tbody>
            </table>
        </div>
        """,
        unsafe_allow_html=True,
    )


def pagina_historico():
    st.markdown("## 📊 Consultar Histórico por Turno")
    df_plantao = load_plantao_df()
    df_ocorrencias = load_ocorrencias_df()

    if df_plantao.empty:
        st.info("Nenhum plantão salvo ainda.")
        return

    df_plantao["km_rodados"] = pd.to_numeric(df_plantao["km_rodados"], errors="coerce")

    with st.container(border=True):
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            datas_disponiveis = sorted(df_plantao["data_plantao"].dropna().unique().tolist(), reverse=True)
            filtro_datas = st.multiselect("Filtrar por Data", options=datas_disponiveis, format_func=format_data_br)
        with col_f2:
            filtro_turno = st.multiselect("Filtrar por Turno", options=["DIA", "NOITE"])
        with col_f3:
            motoristas_disponiveis = sorted(df_plantao["colaborador"].dropna().unique().tolist())
            filtro_motorista = st.multiselect("Filtrar por Motorista", options=motoristas_disponiveis)

    df_filtrado = df_plantao.copy()
    if filtro_datas:
        df_filtrado = df_filtrado[df_filtrado["data_plantao"].isin(filtro_datas)]
    if filtro_turno:
        df_filtrado = df_filtrado[df_filtrado["turno"].isin(filtro_turno)]
    if filtro_motorista:
        df_filtrado = df_filtrado[df_filtrado["colaborador"].isin(filtro_motorista)]

    if df_filtrado.empty:
        st.warning("Nenhum registro encontrado.")
        return

    ordem_turno = {"DIA": 0, "NOITE": 1}
    chaves = df_filtrado[["data_plantao", "turno"]].drop_duplicates()
    chaves_ordenadas = sorted(
        chaves.itertuples(index=False, name=None),
        key=lambda k: (k[0], ordem_turno.get(k[1], 2)),
        reverse=True,
    )

    for data_str, turno in chaves_ordenadas:
        bloco = df_filtrado[(df_filtrado["data_plantao"] == data_str) & (df_filtrado["turno"] == turno)]
        _renderizar_bloco_historico(data_str, turno, bloco, df_ocorrencias)

    st.divider()
    excel_bytes = export_consolidated_excel(df_filtrado, df_ocorrencias)
    st.download_button(
        "📥 Exportar Relatório Consolidado para Excel",
        data=excel_bytes,
        file_name=f"relatorio_consolidado_epr_{date.today().isoformat()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )


def main():
    inject_custom_css()
    render_sidebar_brand()

    pagina = st.navigation(
        [
            st.Page(pagina_inicio, title="Início / Painel Geral", icon="🏠", default=True),
            st.Page(pagina_lancar_plantao, title="Lançar Plantão Diário", icon="📝"),
            st.Page(pagina_historico, title="Consultar Histórico por Turno", icon="📊"),
        ]
    )
    pagina.run()
    render_sidebar_footer()


if __name__ == "__main__":
    main()
